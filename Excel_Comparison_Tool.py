import os
import sys
import json
from typing import List, Dict, Tuple

import pandas as pd
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import (
    QApplication, QWidget, QLabel, QPushButton, QFileDialog,
    QVBoxLayout, QHBoxLayout, QComboBox, QListWidget, QListWidgetItem,
    QTextEdit, QMessageBox, QProgressBar, QCheckBox, QLineEdit, QGroupBox,
    QAbstractItemView
)

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


CONFIG_FILE = "excel_compare_config.json"
OUTPUT_FILE = "diff_result.xlsx"


# ----------------------------
# 工具函数
# ----------------------------
def load_config() -> dict:
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_config(cfg: dict):
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def list_sheets(xlsx_path: str) -> List[str]:
    xls = pd.ExcelFile(xlsx_path)
    return list(xls.sheet_names)


def read_sheet_columns(xlsx_path: str, sheet_name: str) -> List[str]:
    df = pd.read_excel(xlsx_path, sheet_name=sheet_name, nrows=0)
    return [str(c) for c in df.columns.tolist()]


def normalize_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    核心：统一类型，避免 1 vs 1.0、NaN 等导致“假差异”
    - 全部转成 pandas StringDtype（保留 NA）
    - 再 fillna("")，确保比较稳定
    """
    df = df.copy()
    for c in df.columns:
        df[c] = df[c].astype("string").fillna("")
    return df


def build_key_series(df: pd.DataFrame, keys: List[str]) -> pd.Series:
    """
    联合主键：用不可见分隔符拼接，保证稳定
    """
    sep = "\u241F"  # ␟
    # df[keys] 已经是 string 且 fillna("")，直接 join
    return df[keys].agg(lambda r: sep.join(r.values.tolist()), axis=1)


def safe_sheet_name(name: str, suffix: str) -> str:
    # Excel sheet 名最大 31 字符
    base = name[:25]
    return (base + "_" + suffix)[:31]


# ----------------------------
# 主程序 GUI
# ----------------------------
class ExcelComparePro(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel 差异对比工具（稳定增强版：双行旧/新 + 黄色高亮）")
        self.resize(920, 680)

        self.cfg = load_config()

        self.old_file = ""
        self.new_file = ""

        self.sheet_list_old: List[str] = []
        self.sheet_list_new: List[str] = []
        self.common_sheets: List[str] = []

        self._init_ui()
        self._load_cfg_into_ui()

    # ----- UI -----
    def _init_ui(self):
        root = QVBoxLayout()

        # 1) 文件
        box_file = QGroupBox("1) 选择文件")
        lay_file = QVBoxLayout()

        r1 = QHBoxLayout()
        self.btn_old = QPushButton("选择旧文件")
        self.btn_old.clicked.connect(self.choose_old)
        self.lbl_old = QLabel("未选择")
        self.lbl_old.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        r1.addWidget(self.btn_old)
        r1.addWidget(self.lbl_old, 1)
        lay_file.addLayout(r1)

        r2 = QHBoxLayout()
        self.btn_new = QPushButton("选择新文件")
        self.btn_new.clicked.connect(self.choose_new)
        self.lbl_new = QLabel("未选择")
        self.lbl_new.setTextInteractionFlags(Qt.TextInteractionFlag.TextSelectableByMouse)
        r2.addWidget(self.btn_new)
        r2.addWidget(self.lbl_new, 1)
        lay_file.addLayout(r2)

        box_file.setLayout(lay_file)
        root.addWidget(box_file)

        # 2) Sheet 选择
        box_sheet = QGroupBox("2) Sheet 选择")
        lay_sheet = QVBoxLayout()

        top = QHBoxLayout()
        self.chk_all_sheets = QCheckBox("对比所有共同 Sheet")
        self.chk_all_sheets.setChecked(True)
        self.chk_all_sheets.stateChanged.connect(self._toggle_sheet_list)

        self.btn_refresh = QPushButton("刷新 Sheet 列表")
        self.btn_refresh.clicked.connect(self.refresh_sheets)

        top.addWidget(self.chk_all_sheets)
        top.addStretch(1)
        top.addWidget(self.btn_refresh)
        lay_sheet.addLayout(top)

        self.list_sheets = QListWidget()
        self.list_sheets.setSelectionMode(QAbstractItemView.SelectionMode.NoSelection)
        self.list_sheets.setEnabled(False)
        lay_sheet.addWidget(self.list_sheets)

        box_sheet.setLayout(lay_sheet)
        root.addWidget(box_sheet)

        # 3) 主键/忽略列
        box_opt = QGroupBox("3) 主键 / 忽略列")
        lay_opt = QVBoxLayout()

        r_sheet_cols = QHBoxLayout()
        r_sheet_cols.addWidget(QLabel("用于读取列名的 Sheet："))
        self.cmb_sheet_for_cols = QComboBox()
        self.cmb_sheet_for_cols.currentTextChanged.connect(self.reload_columns)
        r_sheet_cols.addWidget(self.cmb_sheet_for_cols, 1)
        self.btn_reload_cols = QPushButton("重新加载列")
        self.btn_reload_cols.clicked.connect(self.reload_columns)
        r_sheet_cols.addWidget(self.btn_reload_cols)
        lay_opt.addLayout(r_sheet_cols)

        # 主键多选
        r_keys = QHBoxLayout()
        r_keys.addWidget(QLabel("主键列（可多选联合主键）："))
        self.list_keys = QListWidget()
        self.list_keys.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        self.list_keys.setMaximumHeight(120)
        r_keys.addWidget(self.list_keys, 1)
        lay_opt.addLayout(r_keys)

        # 忽略列多选
        r_ign = QHBoxLayout()
        r_ign.addWidget(QLabel("忽略列（精确列名，多选）："))
        self.list_ignore = QListWidget()
        self.list_ignore.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        self.list_ignore.setMaximumHeight(120)
        r_ign.addWidget(self.list_ignore, 1)
        lay_opt.addLayout(r_ign)

        # 忽略关键词
        r_kw = QHBoxLayout()
        r_kw.addWidget(QLabel("忽略列名包含关键词（逗号分隔）："))
        self.edt_ignore_contains = QLineEdit()
        self.edt_ignore_contains.setPlaceholderText("例如：更新时间,更新人,修改时间,操作人")
        r_kw.addWidget(self.edt_ignore_contains, 1)
        lay_opt.addLayout(r_kw)

        box_opt.setLayout(lay_opt)
        root.addWidget(box_opt)

        # 4) 执行
        box_run = QGroupBox("4) 执行 / 输出")
        lay_run = QVBoxLayout()

        r_run = QHBoxLayout()
        self.btn_run = QPushButton("开始对比并生成 diff_result.xlsx")
        self.btn_run.clicked.connect(self.run_compare)
        self.btn_save = QPushButton("保存配置")
        self.btn_save.clicked.connect(self.save_cfg_from_ui)
        r_run.addWidget(self.btn_run)
        r_run.addWidget(self.btn_save)
        r_run.addStretch(1)
        lay_run.addLayout(r_run)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        lay_run.addWidget(self.progress)

        self.log = QTextEdit()
        self.log.setReadOnly(True)
        lay_run.addWidget(self.log)

        box_run.setLayout(lay_run)
        root.addWidget(box_run)

        self.setLayout(root)

    def _toggle_sheet_list(self):
        self.list_sheets.setEnabled(not self.chk_all_sheets.isChecked())

    def log_msg(self, msg: str):
        self.log.append(msg)
        QApplication.processEvents()

    # ----- 文件 -----
    def choose_old(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择旧文件", "", "Excel (*.xlsx *.xls)")
        if path:
            self.old_file = path
            self.lbl_old.setText(path)
            self.cfg["old_file"] = path
            self._try_refresh_all()

    def choose_new(self):
        path, _ = QFileDialog.getOpenFileName(self, "选择新文件", "", "Excel (*.xlsx *.xls)")
        if path:
            self.new_file = path
            self.lbl_new.setText(path)
            self.cfg["new_file"] = path
            self._try_refresh_all()

    def _try_refresh_all(self):
        if self.old_file and self.new_file:
            self.refresh_sheets()

    # ----- Sheet & 列 -----
    def refresh_sheets(self):
        if not self.old_file or not self.new_file:
            QMessageBox.warning(self, "提示", "请先选择旧文件和新文件")
            return
        try:
            self.sheet_list_old = list_sheets(self.old_file)
            self.sheet_list_new = list_sheets(self.new_file)
            self.common_sheets = [s for s in self.sheet_list_old if s in self.sheet_list_new]
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取 Sheet 失败：\n{e}")
            return

        self.list_sheets.clear()
        for s in self.common_sheets:
            it = QListWidgetItem(s)
            it.setCheckState(Qt.CheckState.Checked)
            self.list_sheets.addItem(it)

        self.cmb_sheet_for_cols.clear()
        self.cmb_sheet_for_cols.addItems(self.common_sheets)

        # 默认 sheet_for_cols
        sheet_for_cols = self.cfg.get("sheet_for_cols")
        if sheet_for_cols and sheet_for_cols in self.common_sheets:
            self.cmb_sheet_for_cols.setCurrentText(sheet_for_cols)

        self.reload_columns()

        self.log_msg(f"共同 Sheet：{len(self.common_sheets)} 个")
        if not self.common_sheets:
            self.log_msg("⚠️ 两个文件没有共同 Sheet 名称，请检查。")

        # 恢复 sheet 勾选
        checks: Dict[str, bool] = self.cfg.get("sheet_checks", {}) or {}
        if checks:
            for i in range(self.list_sheets.count()):
                it = self.list_sheets.item(i)
                if it.text() in checks:
                    it.setCheckState(Qt.CheckState.Checked if checks[it.text()] else Qt.CheckState.Unchecked)

    def reload_columns(self):
        if not self.old_file or not self.new_file:
            return
        sheet = self.cmb_sheet_for_cols.currentText()
        if not sheet:
            return
        try:
            cols_old = read_sheet_columns(self.old_file, sheet)
            cols_new = read_sheet_columns(self.new_file, sheet)
            cols = [c for c in cols_old if c in cols_new]
        except Exception as e:
            QMessageBox.critical(self, "错误", f"读取列失败：\n{e}")
            return

        self.list_keys.clear()
        self.list_ignore.clear()
        for c in cols:
            self.list_keys.addItem(QListWidgetItem(c))
            self.list_ignore.addItem(QListWidgetItem(c))

        self.cfg["sheet_for_cols"] = sheet

        # 恢复上次选择
        self._restore_key_ignore_selection()

    def _restore_key_ignore_selection(self):
        key_cols = set(self.cfg.get("key_cols", []) or [])
        ignore_cols = set(self.cfg.get("ignore_cols", []) or [])
        ignore_contains = self.cfg.get("ignore_contains", "")

        self.edt_ignore_contains.setText(ignore_contains)

        for i in range(self.list_keys.count()):
            it = self.list_keys.item(i)
            if it.text() in key_cols:
                it.setSelected(True)

        for i in range(self.list_ignore.count()):
            it = self.list_ignore.item(i)
            if it.text() in ignore_cols:
                it.setSelected(True)

    def _get_selected_texts(self, lw: QListWidget) -> List[str]:
        return [it.text() for it in lw.selectedItems()]

    def _get_checked_sheets(self) -> List[str]:
        if self.chk_all_sheets.isChecked():
            return self.common_sheets
        sheets = []
        for i in range(self.list_sheets.count()):
            it = self.list_sheets.item(i)
            if it.checkState() == Qt.CheckState.Checked:
                sheets.append(it.text())
        return sheets

    # ----- 配置 -----
    def save_cfg_from_ui(self):
        self.cfg["old_file"] = self.old_file
        self.cfg["new_file"] = self.new_file
        self.cfg["sheet_for_cols"] = self.cmb_sheet_for_cols.currentText()
        self.cfg["key_cols"] = self._get_selected_texts(self.list_keys)
        self.cfg["ignore_cols"] = self._get_selected_texts(self.list_ignore)
        self.cfg["ignore_contains"] = self.edt_ignore_contains.text().strip()
        self.cfg["all_sheets"] = self.chk_all_sheets.isChecked()
        self.cfg["sheet_checks"] = {
            self.list_sheets.item(i).text(): (self.list_sheets.item(i).checkState() == Qt.CheckState.Checked)
            for i in range(self.list_sheets.count())
        }
        save_config(self.cfg)
        QMessageBox.information(self, "提示", f"配置已保存到 {CONFIG_FILE}")

    def _load_cfg_into_ui(self):
        self.old_file = self.cfg.get("old_file", "") or ""
        self.new_file = self.cfg.get("new_file", "") or ""

        if self.old_file:
            self.lbl_old.setText(self.old_file)
        if self.new_file:
            self.lbl_new.setText(self.new_file)

        self.chk_all_sheets.setChecked(bool(self.cfg.get("all_sheets", True)))
        self._toggle_sheet_list()

        if self.old_file and self.new_file and os.path.exists(self.old_file) and os.path.exists(self.new_file):
            self.refresh_sheets()

    # ----------------------------
    # 核心：对比 + 输出（双行旧/新 + 黄色高亮）
    # ----------------------------
    def run_compare(self):
        if not self.old_file or not self.new_file:
            QMessageBox.warning(self, "提示", "请先选择旧文件和新文件")
            return
        if not os.path.exists(self.old_file) or not os.path.exists(self.new_file):
            QMessageBox.warning(self, "提示", "文件路径不存在，请重新选择")
            return

        sheets = self._get_checked_sheets()
        if not sheets:
            QMessageBox.warning(self, "提示", "请选择至少一个 Sheet")
            return

        keys = self._get_selected_texts(self.list_keys)
        if not keys:
            QMessageBox.warning(self, "提示", "请选择主键列（至少1列）")
            return

        ignore_cols = self._get_selected_texts(self.list_ignore)
        ignore_contains = [s.strip() for s in self.edt_ignore_contains.text().split(",") if s.strip()]

        # 自动保存配置
        self.save_cfg_from_ui()

        self.btn_run.setEnabled(False)
        self.progress.setValue(0)
        self.log.clear()
        self.log_msg("开始任务...")

        # Excel 高亮填充
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        # 写 Excel
        try:
            writer = pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl")
            summary_rows = []
            total = max(1, len(sheets))

            for idx, sheet in enumerate(sheets, start=1):
                self.log_msg(f"\n=== 对比 Sheet: {sheet} ({idx}/{total}) ===")

                df_old = pd.read_excel(self.old_file, sheet_name=sheet)
                df_new = pd.read_excel(self.new_file, sheet_name=sheet)

                df_old.columns = [str(c) for c in df_old.columns]
                df_new.columns = [str(c) for c in df_new.columns]

                common_cols = [c for c in df_old.columns if c in df_new.columns]
                if not common_cols:
                    self.log_msg(f"Sheet {sheet}：两边没有共同列，跳过。")
                    summary_rows.append([sheet, 0, 0, 0, "无共同列"])
                    self.progress.setValue(int(idx * 100 / total))
                    QApplication.processEvents()
                    continue

                # 主键必须存在于共同列
                for k in keys:
                    if k not in common_cols:
                        raise RuntimeError(f"Sheet【{sheet}】缺少主键列【{k}】（两边必须都有该列）")

                # 忽略列集合：精确 + 包含关键词
                ignore_set = set(ignore_cols or [])
                if ignore_contains:
                    for c in common_cols:
                        for kw in ignore_contains:
                            if kw and kw in c:
                                ignore_set.add(c)

                compare_cols = [c for c in common_cols if c not in ignore_set]
                # 主键强制保留
                for k in keys:
                    if k not in compare_cols:
                        compare_cols.append(k)

                # 只取比较列
                df_old = df_old[compare_cols].copy()
                df_new = df_new[compare_cols].copy()

                # 统一类型
                df_old = normalize_df(df_old)
                df_new = normalize_df(df_new)

                # 构建联合主键
                key_old = build_key_series(df_old, keys)
                key_new = build_key_series(df_new, keys)

                df_old.insert(0, "__KEY__", key_old.values)
                df_new.insert(0, "__KEY__", key_new.values)

                # 主键重复：保留最后一条
                df_old = df_old.drop_duplicates(subset=["__KEY__"], keep="last").set_index("__KEY__", drop=True)
                df_new = df_new.drop_duplicates(subset=["__KEY__"], keep="last").set_index("__KEY__", drop=True)

                # 新增/删除
                added = df_new.loc[~df_new.index.isin(df_old.index)]
                deleted = df_old.loc[~df_old.index.isin(df_new.index)]

                # 修改
                common_idx = df_old.index.intersection(df_new.index)
                old_common = df_old.loc[common_idx]
                new_common = df_new.loc[common_idx]

                non_key_cols = [c for c in compare_cols if c not in keys]
                if non_key_cols:
                    diff_mask = (old_common[non_key_cols] != new_common[non_key_cols]).any(axis=1)
                else:
                    diff_mask = pd.Series([False] * len(common_idx), index=common_idx)

                changed_old = old_common.loc[diff_mask]
                changed_new = new_common.loc[diff_mask]

                # 写新增/删除
                added_name = safe_sheet_name(sheet, "新增")
                deleted_name = safe_sheet_name(sheet, "删除")
                mod_name = safe_sheet_name(sheet, "修改")

                added.to_excel(writer, sheet_name=added_name, index=True)
                deleted.to_excel(writer, sheet_name=deleted_name, index=True)

                # 写“修改”：双行旧/新 + 新行差异黄色
                # 表头：类型 + 原列
                out_cols = ["类型"] + compare_cols
                out_rows: List[List[str]] = []
                highlight_cells: List[Tuple[int, int]] = []  # (row, col) in Excel 1-based

                # Excel 行号：第1行是表头，所以数据从2开始
                excel_row = 2

                for key_val in changed_old.index:
                    o = changed_old.loc[key_val]
                    n = changed_new.loc[key_val]

                    old_row = ["旧"] + [o[c] for c in compare_cols]
                    new_row = ["新"] + [n[c] for c in compare_cols]
                    out_rows.append(old_row)
                    out_rows.append(new_row)

                    # 标记新行差异列黄色（注意：Excel 列从1开始；第1列是“类型”）
                    for j, col in enumerate(compare_cols, start=2):  # j=2 对应 compare_cols第1列
                        if o[col] != n[col]:
                            highlight_cells.append((excel_row + 1, j))  # 新行 = excel_row + 1

                    excel_row += 2

                mod_df = pd.DataFrame(out_rows, columns=out_cols)
                mod_df.to_excel(writer, sheet_name=mod_name, index=False)

                # openpyxl 上色、冻结首行、可选：自动列宽
                ws = writer.book[mod_name]
                ws.freeze_panes = "A2"

                for r, c in highlight_cells:
                    ws.cell(row=r, column=c).fill = yellow_fill

                # 可选：简单自适应列宽（避免太夸张，限制最大宽度）
                for col_idx in range(1, len(out_cols) + 1):
                    col_letter = get_column_letter(col_idx)
                    max_len = 0
                    # 只扫前 200 行避免慢
                    for rr in range(1, min(ws.max_row, 200) + 1):
                        v = ws.cell(rr, col_idx).value
                        if v is None:
                            continue
                        max_len = max(max_len, len(str(v)))
                    ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 40)

                summary_rows.append([sheet, len(added), len(deleted), len(changed_old), "OK"])
                self.log_msg(f"Sheet {sheet}：新增 {len(added)}，删除 {len(deleted)}，修改 {len(changed_old)}")

                self.progress.setValue(int(idx * 100 / total))
                QApplication.processEvents()

            # 汇总
            summary_df = pd.DataFrame(summary_rows, columns=["Sheet", "新增", "删除", "修改", "状态"])
            summary_df.to_excel(writer, sheet_name="汇总", index=False)

            writer.close()
        except Exception as e:
            self.btn_run.setEnabled(True)
            QMessageBox.critical(self, "错误", f"执行失败：\n{e}")
            return

        self.btn_run.setEnabled(True)
        self.progress.setValue(100)
        QMessageBox.information(self, "完成", f"对比完成！已生成：{OUTPUT_FILE}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = ExcelComparePro()
    w.show()
    sys.exit(app.exec())
